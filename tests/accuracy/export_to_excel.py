"""
Excel Export Module
Converts JSON accuracy test results into comprehensive Excel reports
"""
import json
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Required packages not installed.")
    print("Please run: pip install pandas openpyxl")
    sys.exit(1)

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))


class AccuracyExcelExporter:
    """Export accuracy test results to Excel with multiple sheets"""
    
    def __init__(self):
        self.results_dir = Path(__file__).parent / "results"
        self.output_dir = self.results_dir
        
    def export_qna_results(self, json_file: str = None, output_file: str = None):
        """Export QnA accuracy results to Excel."""
        if not json_file:
            json_file = self.results_dir / "qna_accuracy_report.json"
        else:
            json_file = Path(json_file)
            
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = self.output_dir / f"qna_accuracy_report_{timestamp}.xlsx"
        else:
            output_file = Path(output_file)
            
        print(f"\n{'='*80}")
        print(f"EXPORTING QnA RESULTS TO EXCEL")
        print(f"{'='*80}")
        print(f"Input:  {json_file}")
        print(f"Output: {output_file}")
        
        try:
            # Load JSON data
            with open(json_file, 'r') as f:
                data = json.load(f)
            
            # Create Excel writer
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Sheet 1: Summary
                self._create_qna_summary_sheet(data, writer)
                
                # Sheet 2: Test Results Detail
                self._create_qna_details_sheet(data, writer)
                
                # Sheet 3: Metrics Breakdown
                self._create_qna_metrics_sheet(data, writer)
                
                # Sheet 4: Category Analysis
                self._create_qna_category_sheet(data, writer)
                
                # Sheet 5: Performance Analysis
                self._create_qna_performance_sheet(data, writer)
                
                # Sheet 6: Failed Cases
                self._create_qna_failed_cases_sheet(data, writer)
            
            # Apply styling
            self._apply_excel_styling(output_file)
            
            print(f"\n✓ Excel report generated successfully!")
            print(f"  Location: {output_file}")
            print(f"{'='*80}\n")
            
            return str(output_file)
            
        except Exception as e:
            print(f"\n✗ Error during export: {e}")
            import traceback
            traceback.print_exc()
            raise
    
    def export_pdf_results(self, json_file: str = None, output_file: str = None):
        """Export PDF accuracy results to Excel."""
        if not json_file:
            json_file = self.results_dir / "pdf_accuracy_report.json"
        else:
            json_file = Path(json_file)
            
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = self.output_dir / f"pdf_accuracy_report_{timestamp}.xlsx"
        else:
            output_file = Path(output_file)
            
        print(f"\n{'='*80}")
        print(f"EXPORTING PDF RESULTS TO EXCEL")
        print(f"{'='*80}")
        print(f"Input:  {json_file}")
        print(f"Output: {output_file}")
        
        try:
            # Load JSON data
            with open(json_file, 'r') as f:
                data = json.load(f)
            
            # Create Excel writer
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Sheet 1: Summary
                self._create_pdf_summary_sheet(data, writer)
                
                # Sheet 2: Test Results Detail
                self._create_pdf_details_sheet(data, writer)
                
                # Sheet 3: Metrics Breakdown
                self._create_pdf_metrics_sheet(data, writer)
                
                # Sheet 4: Document Analysis
                self._create_pdf_document_sheet(data, writer)
                
                # Sheet 5: Retrieval Analysis
                self._create_pdf_retrieval_sheet(data, writer)
                
                # Sheet 6: Failed Cases
                self._create_pdf_failed_cases_sheet(data, writer)
            
            # Apply styling
            self._apply_excel_styling(output_file)
            
            print(f"\n✓ Excel report generated successfully!")
            print(f"  Location: {output_file}")
            print(f"{'='*80}\n")
            
            return str(output_file)
            
        except Exception as e:
            print(f"\n✗ Error during export: {e}")
            import traceback
            traceback.print_exc()
            raise
    
    def _create_qna_summary_sheet(self, data: Dict[str, Any], writer):
        """Create summary sheet for QnA results"""
        summary_data = []
        
        # Test Information
        summary_data.append(["TEST INFORMATION", ""])
        summary_data.append(["Test Type", data.get("test_type", "N/A")])
        summary_data.append(["Timestamp", data.get("timestamp", "N/A")])
        summary_data.append(["Persona ID", data.get("persona_id", "N/A")])
        summary_data.append(["Dataset ID", data.get("dataset_id", "N/A")])
        summary = data.get("summary", {})
        summary_data.append(["Total Test Cases", summary.get("total_tests", 0)])
        summary_data.append(["", ""])
        
        # Aggregate Metrics
        agg = data.get("aggregate_metrics", {})
        summary_data.append(["AGGREGATE METRICS", ""])
        
        # Handle nested dict structure
        semantic_sim = agg.get('semantic_similarity', {})
        if isinstance(semantic_sim, dict):
            summary_data.append(["Avg Semantic Similarity", f"{semantic_sim.get('mean', 0):.4f}"])
        else:
            summary_data.append(["Avg Semantic Similarity", f"{semantic_sim:.4f}"])
        
        bleu = agg.get('bleu_score', {})
        if isinstance(bleu, dict):
            summary_data.append(["Avg BLEU Score", f"{bleu.get('mean', 0):.4f}"])
        else:
            summary_data.append(["Avg BLEU Score", f"{bleu:.4f}"])
        
        exact_match = agg.get('exact_match_ratio', {})
        if isinstance(exact_match, dict):
            summary_data.append(["Avg Exact Match Ratio", f"{exact_match.get('mean', 0):.4f}"])
        else:
            summary_data.append(["Avg Exact Match Ratio", f"{exact_match:.4f}"])
        
        word_overlap = agg.get('word_overlap_ratio', {})
        if isinstance(word_overlap, dict):
            summary_data.append(["Avg Word Overlap", f"{word_overlap.get('mean', 0):.4f}"])
        else:
            summary_data.append(["Avg Word Overlap", f"{word_overlap:.4f}"])
        
        summary_data.append(["", ""])
        
        # Retrieval Metrics
        summary_data.append(["RETRIEVAL METRICS", ""])
        
        retrieval_acc = agg.get('retrieval_accuracy', {})
        if isinstance(retrieval_acc, dict):
            summary_data.append(["Retrieval Accuracy", f"{retrieval_acc.get('accuracy', 0):.4f}"])
            summary_data.append(["Correct Retrievals", retrieval_acc.get('correct_retrievals', 0)])
            summary_data.append(["Total Tests", retrieval_acc.get('total_with_expected_faq', 0)])
        else:
            summary_data.append(["Retrieval Accuracy", f"{retrieval_acc:.4f}"])
        
        retrieval_rank = agg.get('retrieval_rank', {})
        if isinstance(retrieval_rank, dict):
            summary_data.append(["Avg Retrieval Rank", f"{retrieval_rank.get('mean', 0):.2f}"])
        else:
            summary_data.append(["Avg Retrieval Rank", f"{retrieval_rank:.2f}"])
        
        summary_data.append(["", ""])
        
        # Performance Metrics
        summary_data.append(["PERFORMANCE METRICS", ""])
        
        perf = agg.get('performance', {})
        if isinstance(perf, dict):
            summary_data.append(["Avg Response Time (s)", f"{perf.get('avg_response_time_seconds', 0):.2f}"])
            summary_data.append(["Min Response Time (s)", f"{perf.get('min_response_time_seconds', 0):.2f}"])
            summary_data.append(["Max Response Time (s)", f"{perf.get('max_response_time_seconds', 0):.2f}"])
            summary_data.append(["Avg Total Tokens", f"{perf.get('avg_total_tokens', 0):.0f}"])
            summary_data.append(["Avg Prompt Tokens", f"{perf.get('avg_prompt_tokens', 0):.0f}"])
            summary_data.append(["Avg Completion Tokens", f"{perf.get('avg_completion_tokens', 0):.0f}"])
        else:
            summary_data.append(["Avg Response Time (s)", "N/A"])
            summary_data.append(["Min Response Time (s)", "N/A"])
            summary_data.append(["Max Response Time (s)", "N/A"])
            summary_data.append(["Avg Total Tokens", "N/A"])
            summary_data.append(["Avg Prompt Tokens", "N/A"])
            summary_data.append(["Avg Completion Tokens", "N/A"])
        
        df = pd.DataFrame(summary_data, columns=["Metric", "Value"])
        df.to_excel(writer, sheet_name="Summary", index=False)
    
    def _create_qna_details_sheet(self, data: Dict[str, Any], writer):
        """Create detailed test results sheet for QnA"""
        results = data.get("test_results", [])
        
        details_data = []
        for idx, result in enumerate(results, 1):
            test_case = result.get("test_case", {})
            response = result.get("response", {})
            metrics = result.get("metrics", {})
            
            details_data.append({
                "Test #": idx,
                "Question": test_case.get("question", ""),
                "Expected Answer": test_case.get("expected_answer", ""),
                "Generated Answer": response.get("generated_answer", ""),
                "Category": test_case.get("category", ""),
                "Difficulty": test_case.get("difficulty", ""),
                "Semantic Similarity": metrics.get("semantic_similarity", 0),
                "BLEU Score": metrics.get("bleu_score", 0),
                "Exact Match": metrics.get("exact_match_ratio", 0),
                "Word Overlap": metrics.get("word_overlap_ratio", 0),
                "Retrieval Correct": "Yes" if metrics.get("retrieval_correct") else "No",
                "Retrieval Rank": metrics.get("retrieval_rank", "N/A"),
                "Sources Count": response.get("sources_count", 0),
                "Response Time (s)": metrics.get("response_time_seconds", 0),
                "Total Tokens": metrics.get("total_tokens", 0)
            })
        
        df = pd.DataFrame(details_data)
        df.to_excel(writer, sheet_name="Test Details", index=False)
    
    def _create_qna_metrics_sheet(self, data: Dict[str, Any], writer):
        """Create metrics breakdown sheet for QnA"""
        results = data.get("test_results", [])
        
        metrics_data = []
        for idx, result in enumerate(results, 1):
            test_case = result.get("test_case", {})
            metrics = result.get("metrics", {})
            
            metrics_data.append({
                "Test #": idx,
                "Question": test_case.get("question", "")[:50] + "...",
                "Semantic Similarity": metrics.get("semantic_similarity", 0),
                "BLEU": metrics.get("bleu_score", 0),
                "Exact Match": metrics.get("exact_match_ratio", 0),
                "Word Overlap": metrics.get("word_overlap_ratio", 0),
                "Retrieval Confidence": metrics.get("retrieval_confidence", 0),
                "Relevance Score": metrics.get("relevance_score", 0)
            })
        
        df = pd.DataFrame(metrics_data)
        df.to_excel(writer, sheet_name="Metrics Breakdown", index=False)
    
    def _create_qna_category_sheet(self, data: Dict[str, Any], writer):
        """Create category analysis sheet for QnA"""
        results = data.get("test_results", [])
        
        # Group by category
        category_stats = {}
        for result in results:
            test_case = result.get("test_case", {})
            metrics = result.get("metrics", {})
            category = test_case.get("category", "general")
            
            if category not in category_stats:
                category_stats[category] = {
                    "count": 0,
                    "semantic_sum": 0,
                    "bleu_sum": 0,
                    "retrieval_correct": 0
                }
            
            category_stats[category]["count"] += 1
            category_stats[category]["semantic_sum"] += metrics.get("semantic_similarity", 0)
            category_stats[category]["bleu_sum"] += metrics.get("bleu_score", 0)
            if metrics.get("retrieval_correct"):
                category_stats[category]["retrieval_correct"] += 1
        
        # Create category data
        category_data = []
        for category, stats in category_stats.items():
            count = stats["count"]
            category_data.append({
                "Category": category,
                "Test Count": count,
                "Avg Semantic Similarity": stats["semantic_sum"] / count if count > 0 else 0,
                "Avg BLEU Score": stats["bleu_sum"] / count if count > 0 else 0,
                "Retrieval Accuracy": stats["retrieval_correct"] / count if count > 0 else 0,
                "Pass Rate (>0.7)": sum(1 for r in results 
                                       if r.get("test_case", {}).get("category") == category 
                                       and r.get("metrics", {}).get("semantic_similarity", 0) > 0.7) / count if count > 0 else 0
            })
        
        df = pd.DataFrame(category_data)
        df.to_excel(writer, sheet_name="Category Analysis", index=False)
    
    def _create_qna_performance_sheet(self, data: Dict[str, Any], writer):
        """Create performance analysis sheet for QnA"""
        results = data.get("test_results", [])
        
        perf_data = []
        for idx, result in enumerate(results, 1):
            test_case = result.get("test_case", {})
            metrics = result.get("metrics", {})
            
            perf_data.append({
                "Test #": idx,
                "Question": test_case.get("question", "")[:50] + "...",
                "Response Time (s)": metrics.get("response_time_seconds", 0),
                "Processing Time (ms)": metrics.get("processing_time_ms", 0),
                "Retrieval Latency (ms)": metrics.get("retrieval_latency_ms", 0),
                "LLM Latency (ms)": metrics.get("llm_latency_ms", 0),
                "Total Tokens": metrics.get("total_tokens", 0),
                "Prompt Tokens": metrics.get("prompt_tokens", 0),
                "Completion Tokens": metrics.get("completion_tokens", 0),
                "Tokens per Second": metrics.get("completion_tokens", 0) / metrics.get("response_time_seconds", 1) if metrics.get("response_time_seconds", 0) > 0 else 0
            })
        
        df = pd.DataFrame(perf_data)
        df.to_excel(writer, sheet_name="Performance", index=False)
    
    def _create_qna_failed_cases_sheet(self, data: Dict[str, Any], writer):
        """Create failed cases sheet for QnA (semantic similarity < 0.5)"""
        results = data.get("test_results", [])
        
        failed_data = []
        for idx, result in enumerate(results, 1):
            test_case = result.get("test_case", {})
            response = result.get("response", {})
            metrics = result.get("metrics", {})
            
            semantic_sim = metrics.get("semantic_similarity", 0)
            if semantic_sim < 0.5:
                failed_data.append({
                    "Test #": idx,
                    "Question": test_case.get("question", ""),
                    "Expected Answer": test_case.get("expected_answer", ""),
                    "Generated Answer": response.get("generated_answer", ""),
                    "Semantic Similarity": semantic_sim,
                    "BLEU Score": metrics.get("bleu_score", 0),
                    "Retrieval Correct": "Yes" if metrics.get("retrieval_correct") else "No",
                    "Category": test_case.get("category", ""),
                    "Difficulty": test_case.get("difficulty", "")
                })
        
        df = pd.DataFrame(failed_data) if failed_data else pd.DataFrame({"Message": ["No failed cases (all tests above 0.5 threshold)"]})
        df.to_excel(writer, sheet_name="Failed Cases", index=False)
    
    def _create_pdf_summary_sheet(self, data: Dict[str, Any], writer):
        """Create summary sheet for PDF results"""
        summary_data = []
        
        # Test Information
        summary_data.append(["TEST INFORMATION", ""])
        summary_data.append(["Test Type", data.get("test_type", "N/A")])
        summary_data.append(["Timestamp", data.get("timestamp", "N/A")])
        summary_data.append(["Persona ID", data.get("persona_id", "N/A")])
        summary_data.append(["Dataset ID", data.get("dataset_id", "N/A")])
        summary = data.get("summary", {})
        summary_data.append(["Total Test Cases", summary.get("total_tests", 0)])
        summary_data.append(["", ""])
        
        # Aggregate Metrics
        agg = data.get("aggregate_metrics", {})
        summary_data.append(["AGGREGATE METRICS", ""])
        
        # Handle nested dict structure
        answer_quality = agg.get('answer_quality', {})
        semantic_sim = answer_quality.get('semantic_similarity', {}) if isinstance(answer_quality, dict) else {}
        if isinstance(semantic_sim, dict):
            summary_data.append(["Avg Semantic Similarity", f"{semantic_sim.get('mean', 0):.4f}"])
        else:
            summary_data.append(["Avg Semantic Similarity", f"{semantic_sim:.4f}"])
        
        bleu = answer_quality.get('bleu_score', {}) if isinstance(answer_quality, dict) else {}
        if isinstance(bleu, dict):
            summary_data.append(["Avg BLEU Score", f"{bleu.get('mean', 0):.4f}"])
        else:
            summary_data.append(["Avg BLEU Score", f"{bleu:.4f}"])
        
        summary_data.append(["", ""])
        
        # Document Retrieval Metrics
        summary_data.append(["DOCUMENT RETRIEVAL", ""])
        
        retrieval_quality = agg.get('retrieval_quality', {})
        doc_precision = retrieval_quality.get('document_precision', 0) if isinstance(retrieval_quality, dict) else 0
        doc_recall = retrieval_quality.get('document_recall', 0) if isinstance(retrieval_quality, dict) else 0
        doc_f1 = retrieval_quality.get('document_f1', 0) if isinstance(retrieval_quality, dict) else 0
        page_precision = retrieval_quality.get('page_precision', 0) if isinstance(retrieval_quality, dict) else 0
        page_recall = retrieval_quality.get('page_recall', 0) if isinstance(retrieval_quality, dict) else 0
        page_f1 = retrieval_quality.get('page_f1', 0) if isinstance(retrieval_quality, dict) else 0
        
        summary_data.append(["Document Precision", f"{doc_precision:.4f}"])
        summary_data.append(["Document Recall", f"{doc_recall:.4f}"])
        summary_data.append(["Document F1 Score", f"{doc_f1:.4f}"])
        summary_data.append(["Page Precision", f"{page_precision:.4f}"])
        summary_data.append(["Page Recall", f"{page_recall:.4f}"])
        summary_data.append(["Page F1 Score", f"{page_f1:.4f}"])
        
        summary_data.append(["", ""])
        
        # Performance Metrics
        summary_data.append(["PERFORMANCE METRICS", ""])
        
        perf = agg.get('performance', {})
        if isinstance(perf, dict):
            summary_data.append(["Avg Response Time (s)", f"{perf.get('avg_response_time_seconds', 0):.2f}"])
            summary_data.append(["Min Response Time (s)", f"{perf.get('min_response_time_seconds', 0):.2f}"])
            summary_data.append(["Max Response Time (s)", f"{perf.get('max_response_time_seconds', 0):.2f}"])
            summary_data.append(["Avg Total Tokens", f"{perf.get('avg_total_tokens', 0):.0f}"])
        else:
            summary_data.append(["Avg Response Time (s)", "N/A"])
            summary_data.append(["Min Response Time (s)", "N/A"])
            summary_data.append(["Max Response Time (s)", "N/A"])
            summary_data.append(["Avg Total Tokens", "N/A"])
        
        df = pd.DataFrame(summary_data, columns=["Metric", "Value"])
        df.to_excel(writer, sheet_name="Summary", index=False)
    
    def _create_pdf_details_sheet(self, data: Dict[str, Any], writer):
        """Create detailed test results sheet for PDF"""
        results = data.get("test_results", [])
        
        details_data = []
        for idx, result in enumerate(results, 1):
            test_case = result.get("test_case", {})
            response = result.get("response", {})
            metrics = result.get("metrics", {})
            
            # Get document IDs as strings
            expected_docs = test_case.get("expected_document_ids", [])
            retrieved_docs = response.get("retrieved_document_ids", [])
            unique_docs = response.get("unique_documents", 0)
            
            # Get pages
            expected_pages = test_case.get("expected_pages", [])
            retrieved_pages = response.get("retrieved_pages", [])
            unique_pages = response.get("unique_pages", 0)
            
            details_data.append({
                "Test #": idx,
                "Question": test_case.get("question", ""),
                "Expected Answer": test_case.get("expected_answer", ""),
                "Generated Answer": response.get("generated_answer", ""),
                "Category": test_case.get("category", ""),
                "Difficulty": test_case.get("difficulty", ""),
                "Context Type": test_case.get("context_type", ""),
                "Expected Documents": len(expected_docs),
                "Retrieved Documents": len(retrieved_docs),
                "Unique Documents": unique_docs,
                "Expected Pages": str(expected_pages),
                "Retrieved Pages": str(retrieved_pages),
                "Unique Pages": unique_pages,
                "Semantic Similarity": metrics.get("semantic_similarity", 0),
                "BLEU Score": metrics.get("bleu_score", 0),
                "ROUGE-1 F1": metrics.get("rouge_1_f1", 0),
                "ROUGE-2 F1": metrics.get("rouge_2_f1", 0),
                "ROUGE-L F1": metrics.get("rouge_l_f1", 0),
                "Doc Precision": metrics.get("doc_retrieval_precision", 0),
                "Doc Recall": metrics.get("doc_retrieval_recall", 0),
                "Doc F1": metrics.get("doc_retrieval_f1", 0),
                "Page Precision": metrics.get("page_retrieval_precision", 0),
                "Page Recall": metrics.get("page_retrieval_recall", 0),
                "Page F1": metrics.get("page_retrieval_f1", 0),
                "Response Time (s)": metrics.get("response_time_seconds", 0),
                "Total Tokens": metrics.get("total_tokens", 0),
                "Sources Count": response.get("sources_count", 0)
            })
        
        df = pd.DataFrame(details_data)
        df.to_excel(writer, sheet_name="Test Details", index=False)
    
    def _create_pdf_metrics_sheet(self, data: Dict[str, Any], writer):
        """Create metrics breakdown sheet for PDF"""
        results = data.get("test_results", [])
        
        metrics_data = []
        for idx, result in enumerate(results, 1):
            test_case = result.get("test_case", {})
            metrics = result.get("metrics", {})
            
            metrics_data.append({
                "Test #": idx,
                "Question": test_case.get("question", "")[:50] + "...",
                "Difficulty": test_case.get("difficulty", ""),
                "Context Type": test_case.get("context_type", ""),
                "Semantic Similarity": metrics.get("semantic_similarity", 0),
                "BLEU": metrics.get("bleu_score", 0),
                "ROUGE-1": metrics.get("rouge_1_f1", 0),
                "ROUGE-2": metrics.get("rouge_2_f1", 0),
                "ROUGE-L": metrics.get("rouge_l_f1", 0),
                "Doc Precision": metrics.get("doc_retrieval_precision", 0),
                "Doc Recall": metrics.get("doc_retrieval_recall", 0),
                "Doc F1": metrics.get("doc_retrieval_f1", 0),
                "Page Precision": metrics.get("page_retrieval_precision", 0),
                "Page Recall": metrics.get("page_retrieval_recall", 0),
                "Page F1": metrics.get("page_retrieval_f1", 0)
            })
        
        df = pd.DataFrame(metrics_data)
        df.to_excel(writer, sheet_name="Metrics Breakdown", index=False)
    
    def _create_pdf_document_sheet(self, data: Dict[str, Any], writer):
        """Create document analysis sheet for PDF"""
        results = data.get("test_results", [])
        
        doc_data = []
        for idx, result in enumerate(results, 1):
            test_case = result.get("test_case", {})
            response = result.get("response", {})
            metrics = result.get("metrics", {})
            
            # Get document IDs
            expected_doc_ids = test_case.get("expected_document_ids", [])
            retrieved_doc_ids = response.get("retrieved_document_ids", [])
            
            # Calculate correct, missing, and incorrect documents
            expected_set = set(expected_doc_ids)
            retrieved_set = set(retrieved_doc_ids)
            
            correct_docs = expected_set & retrieved_set
            missing_docs = expected_set - retrieved_set
            incorrect_docs = retrieved_set - expected_set
            
            doc_data.append({
                "Test #": idx,
                "Question": test_case.get("question", "")[:50] + "...",
                "Expected Documents": len(expected_doc_ids),
                "Retrieved Documents": len(set(retrieved_doc_ids)),
                "Correct Documents": len(correct_docs),
                "Missing Documents": len(missing_docs),
                "Incorrect Documents": len(incorrect_docs),
                "Document Precision": metrics.get("doc_retrieval_precision", 0),
                "Document Recall": metrics.get("doc_retrieval_recall", 0),
                "Document F1": metrics.get("doc_retrieval_f1", 0),
                "Sources Count": response.get("sources_count", 0),
                "Unique Documents": response.get("unique_documents", 0)
            })
        
        df = pd.DataFrame(doc_data)
        df.to_excel(writer, sheet_name="Document Analysis", index=False)
    
    def _create_pdf_retrieval_sheet(self, data: Dict[str, Any], writer):
        """Create retrieval analysis sheet for PDF"""
        results = data.get("test_results", [])
        
        retrieval_data = []
        for idx, result in enumerate(results, 1):
            test_case = result.get("test_case", {})
            response = result.get("response", {})
            metrics = result.get("metrics", {})
            
            # Get pages
            expected_pages = test_case.get("expected_pages", [])
            retrieved_pages = response.get("retrieved_pages", [])
            
            # Calculate page metrics
            expected_set = set(expected_pages)
            retrieved_set = set(retrieved_pages)
            
            correct_pages = expected_set & retrieved_set
            missing_pages = expected_set - retrieved_set
            incorrect_pages = retrieved_set - expected_set
            
            retrieval_data.append({
                "Test #": idx,
                "Question": test_case.get("question", "")[:50] + "...",
                "Expected Pages": str(sorted(expected_pages)),
                "Retrieved Pages": str(sorted(list(set(retrieved_pages)))),
                "Correct Pages": len(correct_pages),
                "Missing Pages": len(missing_pages),
                "Incorrect Pages": len(incorrect_pages),
                "Page Precision": metrics.get("page_retrieval_precision", 0),
                "Page Recall": metrics.get("page_retrieval_recall", 0),
                "Page F1": metrics.get("page_retrieval_f1", 0),
                "Sources Count": response.get("sources_count", 0),
                "Unique Pages": response.get("unique_pages", 0),
                "Context Type": test_case.get("context_type", "")
            })
        
        df = pd.DataFrame(retrieval_data)
        df.to_excel(writer, sheet_name="Retrieval Analysis", index=False)
    
    def _create_pdf_failed_cases_sheet(self, data: Dict[str, Any], writer):
        """Create failed cases sheet for PDF (semantic similarity < 0.5)"""
        results = data.get("test_results", [])
        
        failed_data = []
        for idx, result in enumerate(results, 1):
            test_case = result.get("test_case", {})
            response = result.get("response", {})
            metrics = result.get("metrics", {})
            
            semantic_sim = metrics.get("semantic_similarity", 0)
            if semantic_sim < 0.5:
                failed_data.append({
                    "Test #": idx,
                    "Question": test_case.get("question", ""),
                    "Expected Answer": test_case.get("expected_answer", ""),
                    "Generated Answer": response.get("generated_answer", ""),
                    "Semantic Similarity": semantic_sim,
                    "BLEU Score": metrics.get("bleu_score", 0),
                    "Doc Precision": metrics.get("doc_retrieval_precision", 0),
                    "Doc Recall": metrics.get("doc_retrieval_recall", 0),
                    "Page Precision": metrics.get("page_retrieval_precision", 0),
                    "Page Recall": metrics.get("page_retrieval_recall", 0),
                    "Expected Docs": len(test_case.get("expected_document_ids", [])),
                    "Expected Pages": len(test_case.get("expected_pages", [])),
                    "Category": test_case.get("category", ""),
                    "Difficulty": test_case.get("difficulty", "")
                })
        
        df = pd.DataFrame(failed_data) if failed_data else pd.DataFrame({"Message": ["No failed cases (all tests above 0.5 threshold)"]})
        df.to_excel(writer, sheet_name="Failed Cases", index=False)
    
    def _apply_excel_styling(self, file_path: Path):
        """Apply professional styling to Excel workbook"""
        wb = load_workbook(file_path)
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Style header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Style section headers (cells with all caps in first column)
            for row in ws.iter_rows(min_row=2):
                if row[0].value and isinstance(row[0].value, str) and row[0].value.isupper() and len(row[0].value.split()) <= 4:
                    row[0].fill = section_fill
                    row[0].font = section_font
                    row[0].alignment = Alignment(horizontal='left', vertical='center')
                else:
                    for cell in row:
                        cell.font = normal_font
                        cell.border = border
                        
                        # Number formatting
                        if cell.value and isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                            if 0 <= cell.value <= 1:
                                cell.number_format = '0.0000'
                            elif cell.value < 100:
                                cell.number_format = '0.00'
                            else:
                                cell.number_format = '#,##0'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = ws['A2']
        
        wb.save(file_path)


def main():
    """Main function to export results"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Export accuracy test results to Excel")
    parser.add_argument(
        "--type",
        choices=["qna", "pdf", "both"],
        default="both",
        help="Type of results to export (default: both)"
    )
    parser.add_argument(
        "--qna-json",
        type=str,
        help="Path to QnA JSON results file"
    )
    parser.add_argument(
        "--pdf-json",
        type=str,
        help="Path to PDF JSON results file"
    )
    parser.add_argument(
        "--qna-output",
        type=str,
        help="Output path for QnA Excel file"
    )
    parser.add_argument(
        "--pdf-output",
        type=str,
        help="Output path for PDF Excel file"
    )
    
    args = parser.parse_args()
    
    exporter = AccuracyExcelExporter()
    
    try:
        if args.type in ["qna", "both"]:
            print("\nExporting QnA results...")
            exporter.export_qna_results(
                json_file=args.qna_json,
                output_file=args.qna_output
            )
        
        if args.type in ["pdf", "both"]:
            print("\nExporting PDF results...")
            exporter.export_pdf_results(
                json_file=args.pdf_json,
                output_file=args.pdf_output
            )
        
        print("\n" + "="*80)
        print("✓ ALL EXPORTS COMPLETED SUCCESSFULLY!")
        print("="*80)
        
    except FileNotFoundError as e:
        print(f"\n✗ Error: {e}")
        print("Please ensure the JSON result files exist before exporting.")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ Error during export: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
